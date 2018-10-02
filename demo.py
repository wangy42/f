import os
from collections import namedtuple
from openpyxl import load_workbook
from functools import reduce

def main():
    scan_for_file()


def scan_for_file():
    list = []
    for item in os.listdir('.'):
        if item.endswith(".xlsx"):
            list.append(item)

    if len(list) == 1:
        read_file(list[0])
    else:
        for index, item in enumerate(list):
            print(str(index) + ": " + item)

        while True:
            x = input("which file to run? ")
            if x.isdigit():
                if int(x) < len(list):
                    read_file(list[int(x)])
                    return
            print("input not valid")


def read_file(filename):
    wb = load_workbook(filename)
    s = wb[wb.sheetnames[0]]
    # header = s['1']
    # l = []
    #
    # for item in header:
    #     if item.value is not None:
    #         l.append(item.value)
    #


    # s = ("\"" + s + "\"")
    # print(s)
    # s = s.lstrip().lstrip('.').lstrip('a')
    # print(s)
    # TestItem = namedtuple("TestItem",s)
    # print(TestItem)
    # s = ', '.join(str(x) for x in l)

    TestItem = namedtuple("TestItem",
                          "Number, Tool, Order, ID, Job, Activity, Index, PRI, Clicks, "
                          "Time, FUN, SIM, PER, STA, SCA, SEC, Comments")
    for r in range(2,10):
        row = s[r]
        l = []
        for i in range(0,17):
            l.append(row[i].value)
        TestItem._make(l)

    print(TestItem)



if __name__ == '__main__':
    main()
